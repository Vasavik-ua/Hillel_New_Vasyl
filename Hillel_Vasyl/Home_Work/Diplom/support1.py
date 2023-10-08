import tkinter as tk
from tkinter import *
import openpyxl
from datetime import date


class Person:
    NAME= []
    SURNAME = []
    SEC_SURNAME = []
    DATE_OF_BIRTH = []
    DATE_OF_DEATH = []
    GENDER = []
    AGE = []


    def column_name(self, shee, array):
        array.insert(0, 'NAME: ')
        for row_id, value in enumerate(array):
            cell = shee.cell(row=row_id + 1, column= 1)
            cell.value = value
    def column_surname(self, shee, array):
        array.insert(0, 'SURNAME: ')
        for row_id, value in enumerate(array):
            cell = shee.cell(row=row_id + 1, column= 2)
            cell.value = value
    def column_sec_surname(self, shee, array):
        array.insert(0, 'SEC SURNAME: ')
        for row_id, value in enumerate(array):
            cell = shee.cell(row=row_id + 1, column= 3)
            cell.value = value
    def column_birth(self, shee, array):
        array.insert(0, 'DATE OF BIRTH: ')
        for row_id, value in enumerate(array):
            cell = shee.cell(row=row_id + 1, column= 4)
            cell.value = value
    def column_death(self, shee, array):
        array.insert(0, 'DATE OF DEATH: ')
        for row_id, value in enumerate(array):
            cell = shee.cell(row=row_id + 1, column= 5)
            cell.value = value

    def column_gender(self, shee, array):
        array.insert(0, 'GENDER: ')
        for row_id, value in enumerate(array):
            cell = shee.cell(row=row_id + 1, column= 6)
            cell.value = value

    def column_age(self, shee, array):
        array.insert(0, 'AGE: ')
        for row_id, value in enumerate(array):
            cell = shee.cell(row=row_id + 1, column= 7)
            cell.value = value





def load_button():
    Person.NAME.append(text_name.get())
    Person.SURNAME.append(text_surname.get())
    Person.SEC_SURNAME.append(text_sec_surname.get())
    Person.DATE_OF_BIRTH.append(text_birth.get())
    Person.DATE_OF_DEATH.append(text_death.get())
    Person.GENDER.append(text_gender.get())


    text_name.delete(0, END)
    text_surname.delete(0, END)
    text_sec_surname.delete(0, END)
    text_birth.delete(0, END)
    text_death.delete(0, END)
    text_gender.delete(0, END)


def check_age(data, other):  # Insert the date
    if other == '':
        dat = date.today()
        dat_t = dat.strftime("%d/%m/%Y")
        sym_v = ''
        for val in dat_t:
            if not val.isdigit():
                sym_v += val
                break
        dat_new = dat_t.split(sym_v)
        sym_val = ''
        for val in data:
            if not val.isdigit():
                sym_val += val
                break
        new_l = data.split(sym_val)
        age = int(dat_new[2]) - int(new_l[2])
        if int(dat_new[1]) <= int(new_l[1]):
            if int(dat_new[0]) < int(new_l[0]):
                age -= 1
        return age
    else:
        sym_v = ''
        for val in other:
            if not val.isdigit():
                sym_v += val
                break
        dat_new = other.split(sym_v)
        sym_val = ''
        for val in data:
            if not val.isdigit():
                sym_val += val
                break
        new_l = data.split(sym_val)
        age = int(dat_new[2]) - int(new_l[2])
        if int(dat_new[1]) <= int(new_l[1]):
            if int(dat_new[0]) < int(new_l[0]):
                age -= 1
        return age



work_book = openpyxl.Workbook()  # Create new file.
work_book.create_sheet(title='Diplom Work', index=0)
sheet = work_book['Diplom Work']


work_book.save('DiploM.xlsx')  # Save new file.


#wb = openpyxl.load_workbook('h.w.20+.xlsx')  #  Open the file.





window = tk.Tk()
window.geometry("700x550")
window.title("!!! DIPLOM !!!")
window.grid_columnconfigure(0, weight=1)

welcome_label = tk.Label(window, text="Register of peoples: ",
                         font=("Helvetica", 13))
welcome_label.grid(row=0, column=1, sticky="W", padx=20, pady=10)

name_label = tk.Label(window, text=" Put the Name: ", font= ("Helvetica", 13))
name_label.grid(row=2, column=0, stick="W", padx=20, pady=10)
text_name = tk.Entry(width=60)
text_name.grid(row=2, column=1, sticky="E", padx=20, pady=10)

surname_label = tk.Label(window, text=" Put the Surname: ", font= ("Helvetica", 13))
surname_label.grid(row=4, column=0, stick="W", padx=20, pady=10)
text_surname = tk.Entry(width=60)
text_surname.grid(row=4, column=1, sticky="E", padx=20, pady=10)

sec_surname_label = tk.Label(window, text=" Put the Second Surname: ", font= ("Helvetica", 13))
sec_surname_label.grid(row=6, column=0, stick="W", padx=20, pady=10)
text_sec_surname = tk.Entry(width=60)
text_sec_surname.grid(row=6, column=1, sticky="E", padx=20, pady=10)

birth_label = tk.Label(window, text=" Put the Birth date: ", font= ("Helvetica", 13))
birth_label.grid(row=8, column=0, stick="W", padx=20, pady=10)
text_birth = tk.Entry(width=60)
text_birth.grid(row=8, column=1, sticky="E", padx=20, pady=10)

death_label = tk.Label(window, text=" Put the Death date: ", font= ("Helvetica", 13))
death_label.grid(row=10, column=0, stick="W", padx=20, pady=10)
text_death = tk.Entry(width=60)
text_death.grid(row=10, column=1, sticky="E", padx=20, pady=10)

gender_label = tk.Label(window, text=" Put the Gender: ", font= ("Helvetica", 13))
gender_label.grid(row=12, column=0, stick="W", padx=20, pady=10)
text_gender = tk.Entry(width=60)
text_gender.grid(row=12, column=1, sticky="E", padx=20, pady=10)


create_button = tk.Button(text="Load People", command=load_button)
create_button.grid(row=16, column= 1, sticky="E", padx=20, pady=10)

create_button = tk.Button(text="Safe File with Input", command=load_button)
create_button.grid(row=22, column= 1, sticky="E", padx=20, pady=10)

create_button = tk.Button(text="Load File ", command=load_button)
create_button.grid(row=26, column= 1, sticky="E", padx=20, pady=10)



print(Person.NAME)

if __name__ == "__main__":
    window.mainloop()
